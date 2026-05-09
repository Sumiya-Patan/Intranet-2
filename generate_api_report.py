from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Data ─────────────────────────────────────────────────────────────────────

ROWS = [
    # LMS
    {
        "module": "LMS",
        "ext_api": "/api/holidays/check?date={date}",
        "http": "GET",
        "tms_caller": "HolidayInfoController",
        "tms_controller": "HolidayInfoController",
        "tms_endpoint": "GET /api/holidays/check/{workDate}",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Direct (RestTemplate in Controller)",
        "intermediary": "—",
        "purpose": "Check if a given date is a public holiday",
    },
    {
        "module": "LMS",
        "ext_api": "/api/holidays/check?date={date}",
        "http": "GET",
        "tms_caller": "TimeSheetController",
        "tms_controller": "TimeSheetController",
        "tms_endpoint": "POST /api/timesheet/create",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Direct (RestTemplate in Controller)",
        "intermediary": "—",
        "purpose": "Validate holiday before allowing timesheet submission",
    },
    {
        "module": "LMS",
        "ext_api": "/api/holidays/month/{month}",
        "http": "GET",
        "tms_caller": "HolidayExcludeUsersService",
        "tms_controller": "HolidayInfoController",
        "tms_endpoint": "GET /api/holidays/currentMonthLeaves",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "HolidayExcludeUsersService",
        "purpose": "Fetch public holidays for the current month",
    },
    {
        "module": "LMS",
        "ext_api": "/api/holidays/month/{month}",
        "http": "GET",
        "tms_caller": "HolidayExcludeUsersService",
        "tms_controller": "HolidayInfoController",
        "tms_endpoint": "GET /api/holidays/currentMonth",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "HolidayExcludeUsersService",
        "purpose": "Fetch public holidays for the current month",
    },
    {
        "module": "LMS",
        "ext_api": "/api/leave-requests/getAllLeaves/{year}/{month}",
        "http": "GET",
        "tms_caller": "LeaveDirectoryService",
        "tms_controller": "HolidayInfoController",
        "tms_endpoint": "GET /api/holidays/currentMonthLeaves",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service (Cached)",
        "intermediary": "LeaveDirectoryService",
        "purpose": "Fetch all leave requests for a given month (cached)",
    },
    {
        "module": "LMS",
        "ext_api": "/api/leave-requests/getLeaveRequests/{userId}?year={year}&month={month}",
        "http": "GET",
        "tms_caller": "LeaveDirectoryService",
        "tms_controller": "HolidayInfoController",
        "tms_endpoint": "GET /api/holidays/currentMonthLeaves",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service (Cached)",
        "intermediary": "LeaveDirectoryService",
        "purpose": "Fetch a specific user's leave requests for a month (cached)",
    },
    {
        "module": "LMS",
        "ext_api": "/api/leave-requests/getAllLeaves/{year}/{month}",
        "http": "GET",
        "tms_caller": "FullHolidayWeekProcessorService",
        "tms_controller": "— (Cron Job)",
        "tms_endpoint": "— (Scheduled Task)",
        "tms_permission": "N/A (System Scheduled)",
        "roles": "System",
        "call_type": "Via Service (Cron)",
        "intermediary": "FullHolidayWeekProcessorService",
        "purpose": "Process full-holiday weeks — fetch leaves for all users",
    },
    # PMS
    {
        "module": "PMS",
        "ext_api": "/projects/tms",
        "http": "GET",
        "tms_caller": "WeeklySummaryService",
        "tms_controller": "WeeklySummaryController",
        "tms_endpoint": "GET /api/timesheet/history",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "WeeklySummaryService",
        "purpose": "Fetch all TMS projects to enrich weekly history with project names",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/tms",
        "http": "GET",
        "tms_caller": "WeeklySummaryService",
        "tms_controller": "WeeklySummaryController",
        "tms_endpoint": "GET /api/timesheet/historyRange",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "WeeklySummaryService",
        "purpose": "Fetch all TMS projects to enrich date-range history with project names",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/tms",
        "http": "GET",
        "tms_caller": "DashboardService",
        "tms_controller": "DashboardController",
        "tms_endpoint": "GET /api/dashboard/summary",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "DashboardService",
        "purpose": "Build projectId→name map for dashboard summary",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/tms",
        "http": "GET",
        "tms_caller": "DashboardService",
        "tms_controller": "DashboardController",
        "tms_endpoint": "GET /api/dashboard/summary/lastMonth",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "DashboardService",
        "purpose": "Build projectId→name map for last-month dashboard summary",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/tms",
        "http": "GET",
        "tms_caller": "DashboardService",
        "tms_controller": "DashboardController",
        "tms_endpoint": "GET /api/dashboard/summary/last3Months",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "DashboardService",
        "purpose": "Build projectId→name map for last-3-months dashboard summary",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/tms",
        "http": "GET",
        "tms_caller": "ProjectDirectoryService",
        "tms_controller": "RMSProjectHoursSummaryController",
        "tms_endpoint": "GET /api/timesheets/RMS/project-hours-summary",
        "tms_permission": "None (no @PreAuthorize)",
        "roles": "Any Authenticated",
        "call_type": "Via Service (Cached)",
        "intermediary": "ProjectDirectoryService",
        "purpose": "Fetch all projects for RMS project hours summary (Spring Cache)",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/tms",
        "http": "GET",
        "tms_caller": "ProjectDirectoryService",
        "tms_controller": "RMSProjectHoursSummaryController",
        "tms_endpoint": "GET /api/timesheets/RMS/project-hours-summary/{projectId}",
        "tms_permission": "None (no @PreAuthorize)",
        "roles": "Any Authenticated",
        "call_type": "Via Service (Cached)",
        "intermediary": "ProjectDirectoryService",
        "purpose": "Fetch all projects for single-project detail (Spring Cache)",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/owner",
        "http": "GET",
        "tms_caller": "ManagerWeeklySummaryService",
        "tms_controller": "ManagerWeeklySummaryController",
        "tms_endpoint": "GET /api/timesheets/manager",
        "tms_permission": "APPROVE_TIMESHEET",
        "roles": "Manager",
        "call_type": "Via Service",
        "intermediary": "ManagerWeeklySummaryService",
        "purpose": "Fetch projects owned by the manager for weekly timesheet view",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/owner",
        "http": "GET",
        "tms_caller": "ManagerWeeklySummaryService",
        "tms_controller": "ManagerWeeklySummaryController",
        "tms_endpoint": "GET /api/timesheets/manager/previous-month/pending",
        "tms_permission": "APPROVE_TIMESHEET",
        "roles": "Manager",
        "call_type": "Via Service",
        "intermediary": "ManagerWeeklySummaryService",
        "purpose": "Fetch manager-owned projects for previous month pending view",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/owner",
        "http": "GET",
        "tms_caller": "ManagerSummaryService",
        "tms_controller": "ManagerSummaryController",
        "tms_endpoint": "GET /api/manager/summary",
        "tms_permission": "APPROVE_TIMESHEET",
        "roles": "Manager",
        "call_type": "Via Service",
        "intermediary": "ManagerSummaryService",
        "purpose": "Fetch manager's owned projects for dashboard summary",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/owner",
        "http": "GET",
        "tms_caller": "ManagerTimeSheetController",
        "tms_controller": "ManagerTimeSheetController",
        "tms_endpoint": "GET /api/manager/users",
        "tms_permission": "APPROVE_TIMESHEET",
        "roles": "Manager",
        "call_type": "Direct (RestTemplate in Controller)",
        "intermediary": "—",
        "purpose": "Get users assigned to manager's projects",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/owner/period?month={m}&year={y}",
        "http": "GET",
        "tms_caller": "ManagerWeeklySummaryService",
        "tms_controller": "ManagerViewReportController",
        "tms_endpoint": "GET /api/report/managerMonthly",
        "tms_permission": "APPROVE_TIMESHEET",
        "roles": "Manager",
        "call_type": "Via Service",
        "intermediary": "ManagerWeeklySummaryService",
        "purpose": "Fetch manager's projects for a specific month/year period report",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/owner/period?month={m}&year={y}",
        "http": "GET",
        "tms_caller": "ManagerMonthlyReportService",
        "tms_controller": "ManagerMonthlyReportPdfController",
        "tms_endpoint": "GET /api/report/managerMonthlyPdf",
        "tms_permission": "APPROVE_TIMESHEET OR VIEW_TIMESHEET",
        "roles": "Manager, Viewer",
        "call_type": "Via Service",
        "intermediary": "ManagerMonthlyReportService",
        "purpose": "Fetch manager's projects to generate monthly PDF report",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/member/{userId}",
        "http": "GET",
        "tms_caller": "TimeSheetService",
        "tms_controller": "ProjectManagementController",
        "tms_endpoint": "GET /api/project-info/managers",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "TimeSheetService",
        "purpose": "Fetch projects where the user is a member, to map manager assignments",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/projects-tasks",
        "http": "GET",
        "tms_caller": "TimeSheetService",
        "tms_controller": "ProjectManagementController",
        "tms_endpoint": "GET /api/project-info/all",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "TimeSheetService",
        "purpose": "Fetch all projects and their tasks for the full project-task view",
    },
    {
        "module": "PMS",
        "ext_api": "/projects/projects-tasks",
        "http": "GET",
        "tms_caller": "RMSProjectHoursSummaryService",
        "tms_controller": "RMSProjectHoursSummaryController",
        "tms_endpoint": "GET /api/timesheets/RMS/project-hours-summary/{projectId}",
        "tms_permission": "None (no @PreAuthorize)",
        "roles": "Any Authenticated",
        "call_type": "Via Service",
        "intermediary": "RMSProjectHoursSummaryService",
        "purpose": "Fetch project tasks from PMS for single-project RMS detail",
    },
    {
        "module": "PMS",
        "ext_api": "/tasks/user/{userId}/tasks",
        "http": "GET",
        "tms_caller": "TimeSheetService",
        "tms_controller": "ProjectManagementController",
        "tms_endpoint": "GET /api/project-info",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "TimeSheetService",
        "purpose": "Fetch tasks assigned to the current user for timesheet entry",
    },
    # UMS
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "ManagerWeeklySummaryService",
        "tms_controller": "ManagerWeeklySummaryController",
        "tms_endpoint": "GET /api/timesheets/manager",
        "tms_permission": "APPROVE_TIMESHEET",
        "roles": "Manager",
        "call_type": "Via Service",
        "intermediary": "ManagerWeeklySummaryService",
        "purpose": "Fetch all users to map user IDs to names in weekly timesheet view",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "ManagerWeeklySummaryService",
        "tms_controller": "ManagerWeeklySummaryController",
        "tms_endpoint": "GET /api/timesheets/manager/previous-month/pending",
        "tms_permission": "APPROVE_TIMESHEET",
        "roles": "Manager",
        "call_type": "Via Service",
        "intermediary": "ManagerWeeklySummaryService",
        "purpose": "Fetch all users to map user IDs for previous month pending view",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "ManagerSummaryService",
        "tms_controller": "ManagerSummaryController",
        "tms_endpoint": "GET /api/manager/summary",
        "tms_permission": "APPROVE_TIMESHEET",
        "roles": "Manager",
        "call_type": "Via Service",
        "intermediary": "ManagerSummaryService",
        "purpose": "Fetch all users to enrich manager summary with user details",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "InternalWeeklySummaryService",
        "tms_controller": "InternalWeeklySummaryController",
        "tms_endpoint": "GET /api/timesheets/internal/summary",
        "tms_permission": "REVIEW_INTERNAL_TIMESHEET OR TIMESHEET_ADMIN",
        "roles": "Supervisor, Admin",
        "call_type": "Via Service",
        "intermediary": "InternalWeeklySummaryService",
        "purpose": "Fetch all users for internal project weekly summary",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "InternalWeeklySummaryService",
        "tms_controller": "InternalWeeklySummaryController",
        "tms_endpoint": "GET /api/timesheets/internal/summary/reportingManager",
        "tms_permission": "REVIEW_INTERNAL_TIMESHEET OR TIMESHEET_ADMIN",
        "roles": "Supervisor, Admin",
        "call_type": "Via Service",
        "intermediary": "InternalWeeklySummaryService",
        "purpose": "Fetch all users for reporting manager's internal summary",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "ManagerMonthlyReportService",
        "tms_controller": "ManagerMonthlyReportPdfController",
        "tms_endpoint": "GET /api/report/managerMonthlyPdf",
        "tms_permission": "APPROVE_TIMESHEET OR VIEW_TIMESHEET",
        "roles": "Manager, Viewer",
        "call_type": "Via Service",
        "intermediary": "ManagerMonthlyReportService",
        "purpose": "Load user names and emails for manager monthly PDF report",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "UserDirectoryService",
        "tms_controller": "UserContoller",
        "tms_endpoint": "GET /api/users",
        "tms_permission": "TIMESHEET_ADMIN",
        "roles": "Admin",
        "call_type": "Via Service (Cached)",
        "intermediary": "UserDirectoryService",
        "purpose": "Return full user list to admin (Spring Cache)",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "UserDirectoryService",
        "tms_controller": "UserContoller",
        "tms_endpoint": "GET /api/users/hours",
        "tms_permission": "⚠️ None (permission commented out)",
        "roles": "⚠️ Any Authenticated (no guard)",
        "call_type": "Via Service (Cached)",
        "intermediary": "UserDirectoryService",
        "purpose": "Fetch users to compute billable/non-billable hours per user",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "HolidayExcludeUsersService",
        "tms_controller": "HolidayExcludeUsersController",
        "tms_endpoint": "GET /api/holiday-exclude-users/allusers",
        "tms_permission": "TIMESHEET_ADMIN",
        "roles": "Admin",
        "call_type": "Via Service",
        "intermediary": "HolidayExcludeUsersService",
        "purpose": "Get all users from UMS to populate holiday exclusion user list",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users?page=1&limit=500",
        "http": "GET",
        "tms_caller": "FullHolidayWeekProcessorService",
        "tms_controller": "— (Cron Job)",
        "tms_endpoint": "— (Scheduled Task)",
        "tms_permission": "N/A (System Scheduled)",
        "roles": "System",
        "call_type": "Via Service (Cron)",
        "intermediary": "FullHolidayWeekProcessorService",
        "purpose": "Fetch all users to process full-holiday week entries automatically",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users/{userId}",
        "http": "GET",
        "tms_caller": "WeeklySummaryService",
        "tms_controller": "WeeklySummaryController",
        "tms_endpoint": "GET /api/timesheet/history",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "WeeklySummaryService",
        "purpose": "Fetch individual user details (full name) for weekly history",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users/{userId}",
        "http": "GET",
        "tms_caller": "WeeklySummaryService",
        "tms_controller": "WeeklySummaryController",
        "tms_endpoint": "GET /api/timesheet/historyRange",
        "tms_permission": "EDIT_TIMESHEET OR APPROVE_TIMESHEET",
        "roles": "Employee, Manager",
        "call_type": "Via Service",
        "intermediary": "WeeklySummaryService",
        "purpose": "Fetch individual user details (full name) for date-range history",
    },
    {
        "module": "UMS",
        "ext_api": "/admin/users/employee/ids",
        "http": "POST",
        "tms_caller": "InternalWeeklySummaryService",
        "tms_controller": "InternalWeeklySummaryController",
        "tms_endpoint": "GET /api/timesheets/internal/summary/reportingManager",
        "tms_permission": "REVIEW_INTERNAL_TIMESHEET OR TIMESHEET_ADMIN",
        "roles": "Supervisor, Admin",
        "call_type": "Via Service",
        "intermediary": "InternalWeeklySummaryService",
        "purpose": "Map EOS employee IDs to UMS user IDs for reporting manager",
    },
    {
        "module": "UMS",
        "ext_api": "/auth/login",
        "http": "POST",
        "tms_caller": "UmsAuthService",
        "tms_controller": "— (Cron Job)",
        "tms_endpoint": "— (Scheduled Task)",
        "tms_permission": "N/A (System Scheduled)",
        "roles": "System",
        "call_type": "Via Service (Cron Auth)",
        "intermediary": "UmsAuthService",
        "purpose": "Authenticate TMS service account to get Bearer token for cron jobs",
    },
    {
        "module": "UMS",
        "ext_api": "/resource/get-all-resources",
        "http": "GET",
        "tms_caller": "RMSTimeSheetService",
        "tms_controller": "RMSTimeSheetController",
        "tms_endpoint": "GET /api/timesheets/RMS/summary",
        "tms_permission": "None (no @PreAuthorize)",
        "roles": "Any Authenticated",
        "call_type": "Via Service",
        "intermediary": "RMSTimeSheetService",
        "purpose": "Fetch resource names for utilization intelligence summary",
    },
    {
        "module": "UMS",
        "ext_api": "/resource/get-all-resources",
        "http": "GET",
        "tms_caller": "RMSTimeSheetService",
        "tms_controller": "RMSTimeSheetController",
        "tms_endpoint": "GET /api/timesheets/RMS/resource-summaries",
        "tms_permission": "None (no @PreAuthorize)",
        "roles": "Any Authenticated",
        "call_type": "Via Service",
        "intermediary": "RMSTimeSheetService",
        "purpose": "Fetch resource names/roles for resource-level summaries",
    },
    {
        "module": "UMS",
        "ext_api": "/resource/get-all-resources",
        "http": "GET",
        "tms_caller": "RMSTimeSheetService",
        "tms_controller": "RMSTimeSheetController",
        "tms_endpoint": "GET /api/timesheets/RMS/users",
        "tms_permission": "None (no @PreAuthorize)",
        "roles": "Any Authenticated",
        "call_type": "Via Service",
        "intermediary": "RMSTimeSheetService",
        "purpose": "Fetch resource roles for simplified user summaries",
    },
    {
        "module": "UMS",
        "ext_api": "/resource/get-all-resources",
        "http": "GET",
        "tms_caller": "RMSTimeSheetService",
        "tms_controller": "RMSTimeSheetController",
        "tms_endpoint": "GET /api/timesheets/RMS/monthly-summary",
        "tms_permission": "None (no @PreAuthorize)",
        "roles": "Any Authenticated",
        "call_type": "Via Service",
        "intermediary": "RMSTimeSheetService",
        "purpose": "Fetch resource names/roles for monthly KPI summary",
    },
    # EOS
    {
        "module": "EOS",
        "ext_api": "/hr/reporting-manager/{managerUuid}/employees",
        "http": "GET",
        "tms_caller": "InternalWeeklySummaryService",
        "tms_controller": "InternalWeeklySummaryController",
        "tms_endpoint": "GET /api/timesheets/internal/summary/reportingManager",
        "tms_permission": "REVIEW_INTERNAL_TIMESHEET OR TIMESHEET_ADMIN",
        "roles": "Supervisor, Admin",
        "call_type": "Via Service",
        "intermediary": "InternalWeeklySummaryService",
        "purpose": "Fetch direct reports of a reporting manager by their UMS UUID",
    },
]

# ─── Styles ───────────────────────────────────────────────────────────────────

MODULE_COLORS = {
    "LMS": "FF8C1A",   # orange
    "PMS": "2E7D32",   # dark green
    "UMS": "1565C0",   # dark blue
    "EOS": "6A1B9A",   # purple
}

MODULE_TEXT_COLORS = {
    "LMS": "FFFFFF",
    "PMS": "FFFFFF",
    "UMS": "FFFFFF",
    "EOS": "FFFFFF",
}

MODULE_ROW_BG = {
    "LMS": "FFF3E0",   # light orange
    "PMS": "E8F5E9",   # light green
    "UMS": "E3F2FD",   # light blue
    "EOS": "F3E5F5",   # light purple
}

WARNING_BG = "FFF9C4"   # pale yellow for ⚠️ rows
WARNING_FONT = "E65100"  # dark orange

HEADER_BG = "263238"
HEADER_FONT = "FFFFFF"

THIN = Side(style="thin", color="BDBDBD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def make_font(hex_color, bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size, name="Calibri")


def make_cell(ws, row, col, value, fill=None, font=None, wrap=True, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    cell.border = BORDER
    return cell


# ─── Build Workbook ───────────────────────────────────────────────────────────

wb = Workbook()

# ──── Sheet 1: External API Calls ────────────────────────────────────────────

ws = wb.active
ws.title = "External API Calls"
ws.freeze_panes = "A2"

COLUMNS = [
    ("#",                      5),
    ("External Module",       14),
    ("External API Endpoint", 45),
    ("HTTP Method",           12),
    ("TMS Caller (Class)",    30),
    ("TMS Controller",        32),
    ("TMS Endpoint",          40),
    ("TMS Permission Required", 38),
    ("Roles / Access Level",  28),
    ("Call Type",             30),
    ("Intermediary Service",  28),
    ("Purpose",               50),
]

# Header row
for col_idx, (header, width) in enumerate(COLUMNS, start=1):
    make_cell(
        ws, 1, col_idx, header,
        fill=make_fill(HEADER_BG),
        font=Font(color=HEADER_FONT, bold=True, size=10, name="Calibri"),
        align="center",
    )
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 22

# Data rows
for row_idx, row in enumerate(ROWS, start=2):
    module = row["module"]
    is_warning = "⚠️" in row["tms_permission"] or "⚠️" in row["roles"]

    row_bg = WARNING_BG if is_warning else MODULE_ROW_BG[module]
    row_fill = make_fill(row_bg)
    row_font = Font(
        color=WARNING_FONT if is_warning else "212121",
        size=9,
        name="Calibri",
    )

    # Module badge cell (colored)
    badge_fill = make_fill(MODULE_COLORS[module])
    badge_font = Font(
        color=MODULE_TEXT_COLORS[module],
        bold=True, size=9, name="Calibri",
    )

    values = [
        row_idx - 1,
        row["module"],
        row["ext_api"],
        row["http"],
        row["tms_caller"],
        row["tms_controller"],
        row["tms_endpoint"],
        row["tms_permission"],
        row["roles"],
        row["call_type"],
        row["intermediary"],
        row["purpose"],
    ]

    for col_idx, val in enumerate(values, start=1):
        if col_idx == 1:
            # Row number — grey bg
            make_cell(ws, row_idx, col_idx, val,
                      fill=make_fill("ECEFF1"),
                      font=Font(color="546E7A", size=9, name="Calibri", bold=True),
                      align="center")
        elif col_idx == 2:
            # Module badge
            make_cell(ws, row_idx, col_idx, val,
                      fill=badge_fill, font=badge_font, align="center")
        elif col_idx == 4:
            # HTTP method — color by verb
            http_colors = {"GET": "1565C0", "POST": "2E7D32", "PUT": "E65100", "DELETE": "C62828"}
            http_bg = {"GET": "E3F2FD", "POST": "E8F5E9", "PUT": "FFF3E0", "DELETE": "FFEBEE"}
            make_cell(ws, row_idx, col_idx, val,
                      fill=make_fill(http_bg.get(val, "F5F5F5")),
                      font=Font(color=http_colors.get(val, "212121"), bold=True, size=9, name="Calibri"),
                      align="center")
        else:
            make_cell(ws, row_idx, col_idx, val,
                      fill=row_fill, font=row_font)

    ws.row_dimensions[row_idx].height = 32

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


# ──── Sheet 2: Legend ────────────────────────────────────────────────────────

ws2 = wb.create_sheet("Legend")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 50

# Title
t = ws2.cell(row=1, column=1, value="TMS External API Report — Legend")
t.font = Font(bold=True, size=13, color="263238", name="Calibri")
t.alignment = Alignment(horizontal="left")
ws2.merge_cells("A1:C1")

# ── Color key ──
ws2.cell(row=3, column=1, value="Color Key").font = Font(bold=True, size=10, name="Calibri")
color_rows = [
    ("LMS", MODULE_COLORS["LMS"], MODULE_ROW_BG["LMS"], "Leave Management Service"),
    ("PMS", MODULE_COLORS["PMS"], MODULE_ROW_BG["PMS"], "Project Management Service"),
    ("UMS", MODULE_COLORS["UMS"], MODULE_ROW_BG["UMS"], "User Management Service"),
    ("EOS", MODULE_COLORS["EOS"], MODULE_ROW_BG["EOS"], "Employee Organization Service"),
    ("⚠️ Warning", "E65100", WARNING_BG, "Endpoint with missing/commented-out @PreAuthorize"),
]
for i, (label, badge_hex, bg_hex, desc) in enumerate(color_rows, start=4):
    ws2.cell(row=i, column=1, value=label).fill = make_fill(badge_hex)
    ws2.cell(row=i, column=1).font = Font(color="FFFFFF" if label != "⚠️ Warning" else "E65100",
                                           bold=True, size=9, name="Calibri")
    ws2.cell(row=i, column=2, value="Row background").fill = make_fill(bg_hex)
    ws2.cell(row=i, column=2).font = Font(size=9, name="Calibri")
    ws2.cell(row=i, column=3, value=desc).font = Font(size=9, name="Calibri")

# ── Permission map ──
perm_start = 11
ws2.cell(row=perm_start - 1, column=1, value="Permission → Role Mapping").font = Font(bold=True, size=10, name="Calibri")

perm_headers = ["Permission", "Role / User Type", "Description"]
for col, h in enumerate(perm_headers, start=1):
    ws2.cell(row=perm_start, column=col, value=h).font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    ws2.cell(row=perm_start, column=col).fill = make_fill(HEADER_BG)

perms = [
    ("EDIT_TIMESHEET",            "Employee",         "Regular employees — can create, edit, view own timesheets"),
    ("APPROVE_TIMESHEET",         "Manager",          "Project managers — can approve/reject team timesheets"),
    ("REVIEW_INTERNAL_TIMESHEET", "Supervisor",       "Supervisors — can review internal project timesheets"),
    ("TIMESHEET_ADMIN",           "Admin",            "System administrators — full access including settings, users, audit logs"),
    ("VIEW_FINANCE_REPORT",       "Finance",          "Finance team — read-only access to financial reports"),
    ("VIEW_TIMESHEET",            "Viewer/Manager",   "Read-only manager access to monthly report PDF"),
]
for i, (perm, role, desc) in enumerate(perms, start=perm_start + 1):
    ws2.cell(row=i, column=1, value=perm).font = Font(size=9, bold=True, name="Calibri")
    ws2.cell(row=i, column=2, value=role).font = Font(size=9, name="Calibri")
    ws2.cell(row=i, column=3, value=desc).font = Font(size=9, name="Calibri")

# ── Call types ──
ct_start = perm_start + len(perms) + 3
ws2.cell(row=ct_start - 1, column=1, value="Call Type Definitions").font = Font(bold=True, size=10, name="Calibri")
call_types = [
    ("Direct (RestTemplate in Controller)", "External HTTP call is made directly inside the controller method — no service intermediary"),
    ("Via Service",                          "Controller delegates to a service class which makes the external HTTP call"),
    ("Via Service (Cached)",                 "Service uses Spring @Cacheable — external call only fires on cache miss"),
    ("Via Service (Cron)",                   "Called by a scheduled cron job, not by an HTTP request — no JWT context"),
    ("Via Service (Cron Auth)",              "Cron job authenticates to UMS to obtain a Bearer token for system-level calls"),
]
ws2.cell(row=ct_start, column=1, value="Call Type").font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
ws2.cell(row=ct_start, column=1).fill = make_fill(HEADER_BG)
ws2.cell(row=ct_start, column=2, value="Meaning").font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
ws2.cell(row=ct_start, column=2).fill = make_fill(HEADER_BG)
ws2.merge_cells(f"B{ct_start}:C{ct_start}")
for i, (ct, meaning) in enumerate(call_types, start=ct_start + 1):
    ws2.cell(row=i, column=1, value=ct).font = Font(size=9, bold=True, name="Calibri")
    ws2.cell(row=i, column=2, value=meaning)
    ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
    ws2.cell(row=i, column=2).font = Font(size=9, name="Calibri")
    ws2.merge_cells(f"B{i}:C{i}")
    ws2.row_dimensions[i].height = 28

# ──── Sheet 3: Summary ───────────────────────────────────────────────────────

ws3 = wb.create_sheet("Summary")
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 45

t = ws3.cell(row=1, column=1, value="External API Call Summary")
t.font = Font(bold=True, size=13, color="263238", name="Calibri")
ws3.merge_cells("A1:C1")

from collections import Counter

module_counts = Counter(r["module"] for r in ROWS)
http_counts = Counter(r["http"] for r in ROWS)
call_type_counts = Counter(
    "Direct (in Controller)" if "Direct" in r["call_type"] else "Via Service Layer"
    for r in ROWS
)

sections = [
    ("Calls by External Module", module_counts),
    ("Calls by HTTP Method", http_counts),
    ("Direct vs Via Service", call_type_counts),
]

cur_row = 3
for title, counts in sections:
    ws3.cell(row=cur_row, column=1, value=title).font = Font(bold=True, size=10, name="Calibri")
    ws3.cell(row=cur_row, column=1).fill = make_fill(HEADER_BG)
    ws3.cell(row=cur_row, column=1).font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    ws3.cell(row=cur_row, column=2, value="Count").font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    ws3.cell(row=cur_row, column=2).fill = make_fill(HEADER_BG)
    cur_row += 1
    for key, count in sorted(counts.items(), key=lambda x: -x[1]):
        ws3.cell(row=cur_row, column=1, value=key).font = Font(size=9, name="Calibri")
        ws3.cell(row=cur_row, column=2, value=count).font = Font(size=9, name="Calibri")
        cur_row += 1
    cur_row += 1

ws3.cell(row=cur_row, column=1, value="Total External API Calls").font = Font(bold=True, size=10, name="Calibri")
ws3.cell(row=cur_row, column=2, value=len(ROWS)).font = Font(bold=True, size=10, name="Calibri")

# ─── Save ─────────────────────────────────────────────────────────────────────

output_path = r"d:\intranet\Intranet-2\Timesheet_External_API_Report.xlsx"
wb.save(output_path)
print(f"Report generated: {output_path}")
print(f"Total rows: {len(ROWS)}")
